#! /usr/bin/env python

from argparse import ArgumentParser
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
from os import path
from sys import stdout, exit

def text_from_excel(xlsx):
	# read as rows
	wb = load_workbook(filename=xlsx, data_only=True)
	text = u''
	for ws in wb.worksheets:
		rows = []
		for r in ws.iter_rows(min_row=1):
			row = [] 
			for cell in r:
				if cell.value is None: row.append('')
				else:
					if cell.number_format.startswith('0.'):
						try:
							# formatted float value
							number_format, prefix = cell.number_format, '0.'
							if number_format.startswith(prefix): number_format = number_format[len(prefix):]
							n = len(number_format)
							val = '%.*f' % (n, Decimal(cell.value))
						except InvalidOperation:
							val = cell.value
						row.append(str(val))
					else: row.append(str(cell.value))
			rows.append(row)
		# text
		text += f'Worksheet {ws.title}\n'
		text += '\n'.join('\t'.join(row) for row in rows) + '\n'
	# text result
	return text

def process_xlsx():
	parser = ArgumentParser(description='Utility to extract text from xlsx file.')
	parser.add_argument("xlsx", help="path of the xlsx file")
	args = parser.parse_args()

	if not path.exists(args.xlsx):
		print('File {} does not exist.'.format(args.xlsx))
		exit(1)

	text = text_from_excel(args.xlsx)
	output = getattr(stdout, 'buffer', stdout)
	output.write(text.encode('utf-8'))

if __name__ == '__main__':
	process_xlsx()
